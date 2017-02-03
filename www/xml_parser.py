#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import datetime
from zipfile import BadZipFile
import logging
logger = logging.getLogger('volgograd_log.xml_parser')

import xlrd
from openpyxl import load_workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from openpyxl.workbook import Workbook
from query import get_str_date

def get_cell(a, row, col):
	s = a.cell(row=row, column=col).value
	if s is None or s == '':
		return None
	return str(s).strip()

class XLSCellError(Exception):
	def __init__(self, row, column, real, expected):
		self.row = row
		self.column = column
		self.real = real
		self.expected = expected

class XLSCellEqualError(XLSCellError):
	pass

class XLSCellInError(XLSCellError):
	pass

class XLSTypeError(XLSCellError):
	pass

def get_float(a, row, col):
	val = get_cell(a, row, col)
	if val is None or val == '':
		return None
	val = val.replace(',', '.')
	if val == '-':
		return None
	try:
		return float(val)
	except ValueError:
		raise XLSTypeError(row, col, val, float)

def get_int(a, row, col):
	val = get_cell(a, row, col)
	if val is None or val == '-' or val == '':
		return None
	try:
		return int(float(val))
	except ValueError:
		raise XLSTypeError(row, col, val, int)


def ensure_cell_equal(a, row, column, expected, get=get_cell):
	real = get(a, row, column)
	if type(real) == str:
		value = real.lower()
	else:
		value = real
	if type(expected) == str:
		need = expected.lower()
	else:
		need = expected
	if value != need:
		raise XLSCellEqualError(row=row, column=column, real=real,
					expected=expected)

def ensure_cell_contains(a, row, column, expected_vals, get=get_cell):
	value = get(a, row, column)
	if value is None:
		raise XLSCellInError(row=row, column=column, real=value,
				     expected=expected_vals)
	lower = value.lower()
	for exp in expected_vals:
		if exp.lower() not in lower:
			raise XLSCellInError(row=row, column=column, real=value,
					     expected=expected_vals)

def validate_forecast_table_header(a, result):
	# Check begin of the table
	ensure_cell_equal(a, 5, 5, 'котельная')
	ensure_cell_contains(a, 5, 3, ['Расчётный', 'темпера', 'турный', 'график'])
	ensure_cell_contains(a, 5, 7, ['Р', 'газа'])
	ensure_cell_equal(a, 5, 8, 'котлы')
	ensure_cell_contains(a, 5, 13, ['Сетевые', 'насосы'])
	ensure_cell_equal(a, 8, 8, 'всего')
	ensure_cell_contains(a, 8, 9, ['в', 'работе'])
	ensure_cell_equal(a, 9, 9, 'котлов')
	ensure_cell_equal(a, 9, 10, 'горелок')
	ensure_cell_equal(a, 8, 11, 'резерв')
	ensure_cell_equal(a, 8, 12, 'ремонт')
	ensure_cell_contains(a, 5, 13, ['Сетевые', 'насосы'])
	ensure_cell_contains(a, 8, 13, ['в', 'работе'])
	ensure_cell_equal(a, 8, 14, 'резерв')
	ensure_cell_equal(a, 8, 15, 'ремонт')
	ensure_cell_contains(a, 5, 16, ['Среднесуточная', 'температура'])
	ensure_cell_equal(a, 6, 16, 'заданная')
	ensure_cell_equal(a, 6, 18, 'фактическая')
	ensure_cell_contains(a, 7, 16, ['tн', 'срзад'])
	ensure_cell_equal(a, 8, 16, 'T1срзад')
	ensure_cell_equal(a, 8, 17, 'T2срзад')
	ensure_cell_equal(a, 8, 18, 'T1срф')
	ensure_cell_equal(a, 8, 19, 'T2срф')
	ensure_cell_contains(a, 5, 20, ['Температура', 'ночь'])
	ensure_cell_equal(a, 6, 20, 'заданная')
	ensure_cell_contains(a, 6, 22, ['фактическая', 'на 6-00 ч.'])
	ensure_cell_contains(a, 7, 20, ['tн', 'нзад'])
	ensure_cell_equal(a, 7, 21, result['expected_temp_air_night'],
			  get=get_float)
	ensure_cell_equal(a, 8, 20, 'T1зад')
	ensure_cell_equal(a, 8, 21, 'T2зад')
	ensure_cell_equal(a, 8, 22, 'T1ф')
	ensure_cell_equal(a, 8, 23, 'T2ф')
	ensure_cell_contains(a, 5, 24, ['давление', 'в', 'сети'])
	ensure_cell_equal(a, 8, 24, 'Р1')
	ensure_cell_equal(a, 8, 25, 'Р2')
	ensure_cell_contains(a, 5, 26, ['расход', 'сетевой', 'воды', 'т/час'])
	ensure_cell_contains(a, 8, 26, ['расч', 'Gр'])
	ensure_cell_contains(a, 8, 27, ['факт', 'Gф'])
	ensure_cell_contains(a, 5, 28, ['расход', 'подпиточной', 'воды'])
	ensure_cell_contains(a, 8, 28, ['расч', 'Gпр', 'т\\час'])
	ensure_cell_equal(a, 8, 29, 'фактический')
	ensure_cell_contains(a, 9, 29, ['на', '6-00', 'т\\час'])
	ensure_cell_contains(a, 9, 30, ['за', 'сутки', 'т\\сут'])
	ensure_cell_contains(a, 9, 31, ['всего с начала месяца', 'т'])
	ensure_cell_contains(a, 5, 34, ['жёсткость', 'мкг-экв\\л'])
	ensure_cell_contains(a, 5, 35, ['прозрач', 'ность', 'см'])
	ensure_cell_equal(a, 8, 34, 'Ж')

def value_is_complex(value):
	if type(value) != str:
		return False
	value = value.lower()
	for letter in value:
		if letter not in 'комплекс':
			return False
	return True

def parse_xls_with_forecast(a):
	result = {}

	ensure_cell_contains(a, 3, 2, ['Сводные данные о работе котельных'])
	#
	# Form the forecast on the next day
	#

	ensure_cell_contains(a, 1, 2, ['Прогноз на'])
	result['forecast_date']      = get_cell(a, 1, 3)
	result['forecast_weather']   = get_cell(a, 2, 2)
	result['forecast_direction'] = get_cell(a, 2, 3)
	result['forecast_speed']     = get_cell(a, 2, 4)

	ensure_cell_equal(a, 1, 6, 'Т день')
	ensure_cell_equal(a, 2, 6, 'Т ночь')
	result['forecast_temp_day_from']   = get_float(a, 1, 7)
	result['forecast_temp_day_to']     = get_float(a, 1, 8)
	result['forecast_temp_night_from'] = get_float(a, 2, 7)
	result['forecast_temp_night_to']   = get_float(a, 2, 8)
	result['date'] = get_cell(a, 3, 27)

	#
	# Average temperature 
	#

	ensure_cell_contains(a, 1, 27, ['tср.возд.', 'за'])
	ensure_cell_contains(a, 2, 27, ['tср. воды.', 'за'])

	# Dates of temperatures must be equal
	
	ensure_cell_equal(a, 2, 29, get_cell(a, 1, 29))
	ensure_cell_equal(a, 2, 29, result['date'])
	result['temp_average_air']   = get_float(a, 1, 31)
	result['temp_average_water'] = get_float(a, 2, 31)

	ensure_cell_equal(a, 4, 3, 'Задаваемая температура наружного воздуха:')
	ensure_cell_equal(a, 4, 15, 'оС')
	result['expected_temp_air_day']     = get_float(a, 4, 14)
	result['expected_temp_air_night']   = get_float(a, 4, 20)
	result['expected_temp_air_all_day'] = get_float(a, 7, 17)

	#
	# Bolier rooms
	#

	ensure_cell_equal(a, 5, 2, 'РАЙОН')
	i = 12
	end = a.max_row - 2
	ditricts = []

	validate_forecast_table_header(a, result)

	fields = ['boilers_all', 'boilers_in_use',
		  'torchs_in_use', 'boilers_reserve', 'boilers_in_repair',
		  'net_pumps_in_work', 'net_pumps_reserve',
		  'net_pumps_in_repair', 'all_day_expected_temp1',
		  'all_day_expected_temp2', 'all_day_real_temp1',
		  'all_day_real_temp2', 'all_night_expected_temp1',
		  'all_night_expected_temp2', 'all_night_real_temp1',
		  'all_night_real_temp2', 'net_pressure1', 'net_pressure2',
		  'net_water_consum_expected_ph', 'net_water_consum_real_ph',
		  'make_up_water_consum_expected_ph',
		  'make_up_water_consum_real_ph',
		  'make_up_water_consum_real_pd',
		  'make_up_water_consum_real_pm']

	while i <= end:
	
		# read all boiler rooms of one district

		district = get_cell(a, i, 2)
		assert(len(district) > 0)
		cur_district = None
		boiler_rooms = []
		while not cur_district and i <= end:

			# read one boiler room

			boiler_room = {}
			if (len(boiler_rooms) == 0):
				boiler_room['district'] = district
			else:
				boiler_room['district'] = None
			boiler_room['T1']   = get_float(a, i, 3)
			boiler_room['T2']   = get_float(a, i, 4)
			boiler_room['name'] = get_cell(a, i, 5)
			boiler_room['gas_pressure'] = get_float(a, i, 7)
			k = 0
			for j in range(8, 16):
				boiler_room[fields[k]] = get_int(a, i, j)
				k += 1
			for j in range(16, 32):
				boiler_room[fields[k]] = get_float(a, i, j)
				k += 1
			hardness = get_cell(a, i, 34)
			if value_is_complex(hardness):
				hardness = None
			else:
				hardness = get_float(a, i, 34)
			transparency = get_cell(a, i, 35)
			if value_is_complex(transparency):
				transparency = None
			else:
				transparency = get_float(a, i, 35)
			boiler_room['hardness']     = hardness
			boiler_room['transparency'] = transparency

			boiler_rooms.append(boiler_room)

			i += 1
			cur_district = get_cell(a, i, 2)
		ditricts.append({'name': district, 'rooms': boiler_rooms})

	result['districts'] = ditricts
	return result

def parse_xls_without_forecast(a):
	result = {}

	ensure_cell_contains(a, 1, 2, ['Сводные данные о работе котельных'])
	result['date'] = get_cell(a, 1, 18)

	#
	# Bolier rooms
	#

	ensure_cell_equal(a, 2, 2, 'РАЙОН')
	i = 6
	ditricts = []

	fields = ['boilers_all', 'boilers_in_use',
		  'torchs_in_use', 'boilers_reserve', 'boilers_in_repair',
		  'all_day_real_temp1', 'all_day_real_temp2',
		  'all_night_real_temp1', 'all_night_real_temp2',
		  'net_pressure1', 'net_pressure2',
		  'net_water_consum_expected_ph', 'net_water_consum_real_ph',
		  'make_up_water_consum_expected_ph',
		  'make_up_water_consum_real_ph',
		  'make_up_water_consum_real_pd',
		  'make_up_water_consum_real_pm']

	end = False
	while not end:
	
		# read all boiler rooms of one district

		district = get_cell(a, i, 2)
		assert(len(district) > 0)
		cur_district = None
		boiler_rooms = []
		while cur_district is None:
			if not get_cell(a, i, 3):
				end = True
				break
			# read one boiler room

			boiler_room = {}
			if (len(boiler_rooms) == 0):
				boiler_room['district'] = district
			else:
				boiler_room['district'] = None
			boiler_room['T1']   = get_float(a, i, 3)
			boiler_room['T2']   = get_float(a, i, 4)
			boiler_room['name'] = get_cell(a, i, 5)
			boiler_room['gas_pressure'] = get_float(a, i, 6)
			k = 0
			for j in range(7, 12):
				boiler_room[fields[k]] = get_int(a, i, j)
				k += 1
			for j in range(12, 24):
				boiler_room[fields[k]] = get_float(a, i, j)
				k += 1
			hardness = get_cell(a, i, 27)
			if value_is_complex(hardness):
				hardness = None
			else:
				hardness = get_float(a, i, 27)
			transparency = get_cell(a, i, 28)
			if value_is_complex(transparency):
				transparency = None
			else:
				transparency = get_float(a, i, 28)
			boiler_room['hardness']     = hardness
			boiler_room['transparency'] = transparency

			boiler_rooms.append(boiler_room)

			i += 1
			cur_district = get_cell(a, i, 2)
		ditricts.append({'name': district, 'rooms': boiler_rooms})

	result['districts'] = ditricts
	return result

def open_xls_as_xlsx(filename):
	# First open using xlrd.
	book = xlrd.open_workbook(file_contents=filename.getvalue())
	index = 0
	nrows, ncols = 0, 0
	while nrows * ncols == 0:
		sheet = book.sheet_by_index(index)
		nrows = sheet.nrows
		ncols = sheet.ncols
		index += 1

	# prepare a xlsx sheet
	book1 = Workbook()
	sheet1 = book1.active

	for row in range(0, nrows):
		for col in range(0, ncols):
			cell = sheet.cell(row, col)
			#
			# Excel saves dates as numbers so need to
			# carefully convert such cells in a
			# readable representation.
			#
			if cell.ctype == xlrd.XL_CELL_DATE:
				value = xlrd.xldate_as_tuple(cell.value,
							     book.datemode)
				value = get_str_date(*(value[:3]))
			else:
				value = cell.value
			sheet1.cell(row=row+1, column=col+1).value = value

	return book1

def parse_xls(file):
	wb = None
	try:
		wb = load_workbook(filename=file, data_only=True)
	except BadZipFile:
		logger.warning("Can't open xls, try to use xlrd...")
		wb = open_xls_as_xlsx(file)
		logger.warning('Success')
	a = wb.active
	if 'Сводные данные о работе котельных' in get_cell(a, 1, 2):
		return parse_xls_without_forecast(a)
	else:
		return parse_xls_with_forecast(a)
